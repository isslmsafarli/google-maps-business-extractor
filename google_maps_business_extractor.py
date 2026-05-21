#!/usr/bin/env python3
"""
Standalone Google Maps business extractor.

Runs gosom/google-maps-scraper via Docker, filters out unusable rows, removes
duplicates, and exports clean CSV, JSON, and Excel files.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import tempfile
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import quote_plus

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - handled at runtime with clear error.
    Workbook = None


SOCIAL_DOMAINS = {
    "instagram.com",
    "facebook.com",
    "fb.com",
    "m.facebook.com",
    "linktr.ee",
    "linktree.com",
    "wa.me",
    "api.whatsapp.com",
    "tiktok.com",
    "youtube.com",
    "youtu.be",
    "twitter.com",
    "x.com",
    "linkedin.com",
    "t.me",
    "telegram.me",
    "threads.net",
}

OUTPUT_COLUMNS = [
    "business_name",
    "address",
    "phone",
    "website",
    "rating",
    "review_count",
    "category",
    "google_maps_link",
]

REQUIRED_FIELDS = [
    "business_name",
    "address",
    "phone",
    "website",
    "category",
    "google_maps_link",
]


@dataclass(frozen=True)
class Business:
    business_name: str
    address: str
    phone: str
    website: str
    rating: float | None
    review_count: int | None
    category: str
    google_maps_link: str


def normalize_website(url: str | None) -> str | None:
    """Return normalized domain ('example.com') or None if unusable."""
    if not url or not url.strip():
        return None
    normalized = url.strip().lower()
    normalized = re.sub(r"^https?://", "", normalized)
    normalized = re.sub(r"^www\.", "", normalized)
    normalized = normalized.split("/")[0].split("?")[0].split("#")[0]
    normalized = normalized.strip()
    return normalized or None


def is_social_only(domain: str | None) -> bool:
    """Check if a normalized domain is just a social/profile link."""
    if not domain:
        return True
    return any(domain == social or domain.endswith("." + social) for social in SOCIAL_DOMAINS)


def slugify_query(query: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", query.strip().lower()).strip("_")
    return slug or "google_maps_results"


def first_value(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def parse_float(value: str) -> float | None:
    if not value:
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def parse_int(value: str) -> int | None:
    if not value:
        return None
    cleaned = re.sub(r"[^\d]", "", str(value))
    if not cleaned:
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def build_google_maps_link(row: dict[str, str], business_name: str, address: str) -> str:
    explicit = first_value(
        row,
        "google_maps_link",
        "maps_link",
        "map_link",
        "link",
        "url",
        "place_url",
        "reviews_link",
    )
    if explicit:
        return explicit

    cid = first_value(row, "cid", "place_id", "data_id")
    if cid:
        return f"https://www.google.com/maps?cid={quote_plus(cid)}"

    query = " ".join(part for part in [business_name, address] if part)
    return f"https://www.google.com/maps/search/?api=1&query={quote_plus(query)}" if query else ""


def run_scraper(query: str, work_dir: Path, depth: int = 5, inactivity: str = "3m") -> Path:
    """Run gosom scraper via Docker and return the generated results.csv path."""
    work_dir.mkdir(parents=True, exist_ok=True)
    queries_file = work_dir / "queries.txt"
    results_file = work_dir / "results.csv"

    queries_file.write_text(query + "\n", encoding="utf-8")
    if results_file.exists():
        results_file.unlink()
    results_file.touch()

    cmd = [
        "docker",
        "run",
        "--rm",
        "-v",
        f"{queries_file}:/queries",
        "-v",
        f"{results_file}:/results",
        "gosom/google-maps-scraper",
        "-depth",
        str(depth),
        "-input",
        "/queries",
        "-results",
        "/results",
        "-exit-on-inactivity",
        inactivity,
    ]

    print(f"[extract] running gosom: query={query!r} depth={depth}")
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        stderr_tail = (proc.stderr or "")[-1000:]
        raise RuntimeError(f"gosom Docker run failed with exit code {proc.returncode}\n{stderr_tail}")

    return results_file


def parse_results(results_file: Path, fallback_category: str) -> list[Business]:
    """Parse gosom results.csv into normalized business records."""
    if not results_file.exists() or results_file.stat().st_size == 0:
        return []

    try:
        csv.field_size_limit(sys.maxsize)
    except OverflowError:
        csv.field_size_limit(2**31 - 1)

    businesses: list[Business] = []
    with results_file.open(newline="", encoding="utf-8", errors="replace") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            business_name = first_value(row, "title", "name", "business", "business_name")
            address = first_value(row, "address", "complete_address", "full_address")
            phone = first_value(row, "phone", "phone_number", "telephone")
            website = first_value(row, "website", "site", "url")
            category = first_value(row, "category", "type", "main_category") or fallback_category
            rating = parse_float(first_value(row, "review_rating", "rating", "google_rating"))
            review_count = parse_int(first_value(row, "review_count", "reviews", "number_of_reviews"))
            maps_link = build_google_maps_link(row, business_name, address)

            businesses.append(
                Business(
                    business_name=business_name,
                    address=address,
                    phone=phone,
                    website=website,
                    rating=rating,
                    review_count=review_count,
                    category=category,
                    google_maps_link=maps_link,
                )
            )

    return businesses


def validate_and_filter(rows: Iterable[Business]) -> tuple[list[Business], Counter]:
    exported: list[Business] = []
    skipped: Counter = Counter()
    seen_domains: set[str] = set()

    for row in rows:
        data = asdict(row)
        missing_required = [field for field in REQUIRED_FIELDS if not data.get(field)]
        if missing_required:
            skipped[f"missing_{missing_required[0]}"] += 1
            continue

        domain = normalize_website(row.website)
        if not domain:
            skipped["no_website"] += 1
            continue

        if is_social_only(domain):
            skipped["social_only"] += 1
            continue

        if domain in seen_domains:
            skipped["duplicate_website"] += 1
            continue

        seen_domains.add(domain)
        exported.append(row)

    return exported, skipped


def write_csv(rows: list[Business], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def write_json(rows: list[Business], path: Path) -> None:
    path.write_text(
        json.dumps([asdict(row) for row in rows], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def write_excel(rows: list[Business], path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export. Run: pip install -r requirements.txt")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Businesses"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(OUTPUT_COLUMNS)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        sheet.append([asdict(row)[column] for column in OUTPUT_COLUMNS])

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        values = [str(column_name)]
        values.extend(str(asdict(row)[column_name] or "") for row in rows)
        width = min(max(len(value) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.freeze_panes = "A2"
    workbook.save(path)


def export_files(rows: list[Business], output_dir: Path, query: str) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    base = output_dir / slugify_query(query)
    paths = {
        "csv": base.with_suffix(".csv"),
        "json": base.with_suffix(".json"),
        "excel": base.with_suffix(".xlsx"),
    }
    write_csv(rows, paths["csv"])
    write_json(rows, paths["json"])
    write_excel(rows, paths["excel"])
    return paths


def print_summary(raw_count: int, exported_count: int, skipped: Counter, paths: dict[str, Path]) -> None:
    skipped_count = sum(skipped.values())
    print()
    print("=" * 58)
    print(f"Businesses found:    {raw_count}")
    print(f"Businesses exported: {exported_count}")
    print(f"Businesses skipped:  {skipped_count}")
    for reason, count in sorted(skipped.items()):
        print(f"  - {reason}: {count}")
    print()
    print("Output files:")
    for file_type, path in paths.items():
        print(f"  {file_type.upper()}: {path}")
    print("=" * 58)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export Google Maps businesses to CSV, JSON, and Excel.")
    parser.add_argument("query", help='Search query, e.g. "dentists London"')
    parser.add_argument("--depth", type=int, default=5, help="gosom scrape depth (default: 5)")
    parser.add_argument("--inactivity", default="3m", help="gosom exit-on-inactivity value (default: 3m)")
    parser.add_argument("--output-dir", default="output", help="Directory for exported files (default: output)")
    parser.add_argument(
        "--keep-work-dir",
        action="store_true",
        help="Keep temporary gosom files for debugging.",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir).resolve()

    if args.keep_work_dir:
        work_dir = Path(".gosom_work").resolve()
        results_file = run_scraper(args.query, work_dir, depth=args.depth, inactivity=args.inactivity)
        raw_rows = parse_results(results_file, fallback_category=args.query)
    else:
        with tempfile.TemporaryDirectory(prefix="gmaps_extractor_") as tmp:
            results_file = run_scraper(args.query, Path(tmp), depth=args.depth, inactivity=args.inactivity)
            raw_rows = parse_results(results_file, fallback_category=args.query)

    exported, skipped = validate_and_filter(raw_rows)
    paths = export_files(exported, output_dir, args.query)
    print_summary(len(raw_rows), len(exported), skipped, paths)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
